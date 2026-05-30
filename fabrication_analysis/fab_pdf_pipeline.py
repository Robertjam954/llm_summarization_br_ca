"""
Fab PDF Pipeline — Fabrication Edge Cases (AI_Has_3)
Extends doc_text_eval_pipeline.py with real PDF processing.

Populates stub columns using PyMuPDF + OpenCV:
  - Image quality: laplacian_variance_blur, tenengrad_sharpness,
                   rms_contrast, mean_brightness, skew_angle_deg, resolution_dpi
  - Flags:         is_blurry, is_low_contrast  (bottom-10th-pctile within set)
  - Text:          text_extraction_method, n_pages_extracted,
                   words_per_page_avg, chars_per_page_avg_redacted
  - OCR:           ocr_confidence_avg_pct (stub — requires MS Foundry OCR)

Maps Excel rows → fab_source folders via surgeon initials + patient initials.
MET_DM is absent from fab_source; that row receives None for all PDF metrics.
"""

import re
import os
import csv
import numpy as np
import cv2
import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

FAB_DIR    = os.path.join(os.path.dirname(__file__), "fab_source")
EXCEL_PATH = os.path.join(FAB_DIR, "llm_validation_failure_analysis.xlsx")
SHEET_NAME = "AI_Has_3 "
RENDER_DPI = 150   # render resolution for quality metrics

# ── Feature definitions (same as original pipeline) ──────────────────────────
FEATURE_TRIPLETS = [
    ("lesion_size",                  "G", "H", "I"),
    ("laterality",                   "J", "K", "L"),
    ("lesion_location",              "M", "N", "O"),
    ("calcifications_asymmetry",     "P", "Q", "R"),
    ("additional_enhancement_mri",   "S", "T", "U"),
    ("extent",                       "V", "W", "X"),
    ("accurate_clip_placement",      "Y", "Z", "AA"),
    ("workup_recommendation",        "AB", "AC", "AD"),
    ("lymph_node",                   "AE", "AF", "AG"),
    ("chronology_preserved",         "AH", "AI", "AJ"),
    ("biopsy_method",                "AK", "AL", "AM"),
    ("invasive_component_size_path", "AN", "AO", "AP"),
    ("histologic_diagnosis",         "AQ", "AR", "AS"),
    ("receptor_status",              "AT", "AU", "AV"),
]
IMAGING_FEATURES = {
    "lesion_size", "laterality", "lesion_location",
    "calcifications_asymmetry", "additional_enhancement_mri",
    "extent", "accurate_clip_placement", "workup_recommendation",
    "lymph_node", "chronology_preserved",
}
PATHOLOGY_FEATURES = {
    "biopsy_method", "invasive_component_size_path",
    "histologic_diagnosis", "receptor_status",
}
QUALITY_KEYWORDS_CONFIRMED = [
    "blurry", "blur", "poor quality pdf", "poor quality",
    "not clear", "low quality", "image quality", "scanned",
]
QUALITY_KEYWORDS_POSSIBLE = [
    "uploaded text", "discrepancy", "text", "long", "truncated",
    "cut off", "missing", "difficult", "unclear", "confused", "confusion",
]
SIZE_ERROR_PATTERN = re.compile(
    r"\d+\.?\d*\s*(cm|mm|instead|rather|than|vs\.?)", re.IGNORECASE
)


# ── Utility helpers ───────────────────────────────────────────────────────────

def col_letter_to_idx(letter: str) -> int:
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result - 1


def surgeon_to_code(surgeon_str: str) -> str:
    """
    'Barrio, Andrea'  → 'AB'
    'El-Tamer, Mahmoud' → 'MET'
    'Tracy-Ann, Moo'  → 'TM'  (handles hyphenated first names)
    """
    surgeon_str = surgeon_str.strip()
    if "," in surgeon_str:
        last, first = surgeon_str.split(",", 1)
    else:
        parts = surgeon_str.split()
        last, first = parts[-1], " ".join(parts[:-1])
    last  = last.strip()
    first = first.strip()
    first_initial = first[0]                            # e.g. 'A' from 'Andrea'
    last_parts = re.split(r"[-\s]+", last)             # split on hyphen or space
    last_code = "".join(p[0] for p in last_parts if p) # e.g. 'ET' from 'El-Tamer'
    return (first_initial + last_code).upper()


def build_folder_map() -> dict:
    """Return {(surgeon_code, patient_initials): abs_folder_path} for fab_source dirs."""
    folder_map = {}
    for name in os.listdir(FAB_DIR):
        full = os.path.join(FAB_DIR, name)
        if not os.path.isdir(full):
            continue
        parts = name.split("_")
        if len(parts) < 2:
            continue
        folder_map[(parts[0], parts[1])] = full
    return folder_map


# ── Comment-based quality classifiers (from original pipeline) ────────────────

def classify_doc_quality(comment: str) -> tuple:
    if not comment:
        return 0, ""
    cl = comment.lower()
    for kw in QUALITY_KEYWORDS_CONFIRMED:
        if kw in cl:
            return 2, f"Confirmed quality issue: '{kw}' in comment"
    for kw in QUALITY_KEYWORDS_POSSIBLE:
        if kw in cl:
            return 1, f"Possible quality concern: '{kw}' in comment"
    if SIZE_ERROR_PATTERN.search(comment):
        return 1, "Possible OCR/text extraction error: numeric size discrepancy"
    return 0, ""


def classify_text_quality(comment: str, doc_quality_flag: int) -> int:
    if not comment:
        return doc_quality_flag
    cl = comment.lower()
    numeric_swap = re.search(
        r"(\d+\.?\d*)\s*(instead of|rather than|was|vs\.?)\s*(\d+\.?\d*)", cl
    )
    if numeric_swap:
        return max(doc_quality_flag, 1)
    return doc_quality_flag


def infer_pdf_type(comment: str, doc_quality_flag: int) -> str:
    if not comment:
        return "unknown"
    cl = comment.lower()
    if any(k in cl for k in ["blurry", "poor quality pdf", "scanned", "image quality"]):
        return "scanned"
    if any(k in cl for k in ["selectable", "native", "word", "export"]):
        return "native"
    if doc_quality_flag == 2:
        return "scanned"
    return "unknown"


def infer_doc_type(comment: str, error_features: list) -> str:
    has_imaging = any(f in IMAGING_FEATURES for f in error_features)
    has_path    = any(f in PATHOLOGY_FEATURES for f in error_features)
    if not error_features:
        return "mixed"
    if has_imaging and has_path:
        return "mixed"
    if has_imaging:
        return "radiology"
    if has_path:
        return "pathology"
    return "mixed"


def quality_attribution(doc_quality_flag: int, text_quality_flag: int,
                        comment: str) -> str:
    if doc_quality_flag == 2 or text_quality_flag == 2:
        return "Confirmed"
    if doc_quality_flag == 1 or text_quality_flag == 1:
        return "Possible"
    if comment and SIZE_ERROR_PATTERN.search(comment):
        return "Possible"
    return "No"


# ── PDF processing helpers ────────────────────────────────────────────────────

def _render_gray(page, dpi: int) -> np.ndarray:
    mat = fitz.Matrix(dpi / 72, dpi / 72)
    pix = page.get_pixmap(matrix=mat, colorspace=fitz.csGRAY)
    return np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width)


def _estimate_skew(gray: np.ndarray) -> float:
    _, binary = cv2.threshold(gray, 0, 255,
                              cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    edges = cv2.Canny(binary, 50, 150, apertureSize=3)
    lines = cv2.HoughLinesP(edges, 1, np.pi / 180,
                            threshold=80, minLineLength=80, maxLineGap=10)
    if lines is None:
        return 0.0
    angles = []
    for ln in lines:
        x1, y1, x2, y2 = ln[0]
        if x2 != x1:
            a = np.degrees(np.arctan2(y2 - y1, x2 - x1))
            if abs(a) < 45:
                angles.append(a)
    return float(np.median(angles)) if angles else 0.0


def _page_embedded_dpi(page) -> float | None:
    """Estimate scan DPI from largest embedded image on the page."""
    images = page.get_images(full=True)
    best = None
    best_px = 0
    for img in images:
        w, h = img[2], img[3]
        if w * h > best_px:
            best_px = w * h
            best = (w, h)
    if best is None or page.rect.width == 0:
        return None
    dpi_x = best[0] / (page.rect.width  / 72)
    dpi_y = best[1] / (page.rect.height / 72)
    return (dpi_x + dpi_y) / 2


def _is_native_pdf(doc) -> bool:
    """Return True if PDF has substantial selectable text (not image-only/scanned)."""
    total_chars = sum(len(p.get_text().strip()) for p in doc)
    return (total_chars / max(len(doc), 1)) > 50


def _page_metrics(page, dpi: int) -> dict:
    gray = _render_gray(page, dpi)
    lap  = cv2.Laplacian(gray, cv2.CV_64F)
    sx   = cv2.Sobel(gray, cv2.CV_64F, 1, 0, ksize=3)
    sy   = cv2.Sobel(gray, cv2.CV_64F, 0, 1, ksize=3)
    return {
        "laplacian_var": float(lap.var()),
        "tenengrad":     float(np.sum(sx ** 2 + sy ** 2)),
        "rms_contrast":  float(np.percentile(gray, 95) - np.percentile(gray, 5)),
        "brightness":    float(np.mean(gray)),
        "skew":          _estimate_skew(gray),
    }


def process_case_pdfs(folder_path: str, dpi: int = RENDER_DPI) -> dict:
    """
    Process all PDFs in a case folder.
    Returns dict of aggregate metrics, or empty dict on total failure.
    """
    pdf_files = sorted(f for f in os.listdir(folder_path) if f.lower().endswith(".pdf"))
    if not pdf_files:
        return {}

    page_metrics = []
    total_pages  = 0
    total_words  = 0
    total_chars  = 0
    native_docs  = 0
    scanned_docs = 0
    dpi_samples  = []

    for fname in pdf_files:
        path = os.path.join(folder_path, fname)
        try:
            doc = fitz.open(path)
        except Exception as exc:
            print(f"  [WARN] Cannot open {fname}: {exc}")
            continue

        is_nat = _is_native_pdf(doc)
        if is_nat:
            native_docs += 1
        else:
            scanned_docs += 1

        for page in doc:
            total_pages += 1
            text = page.get_text()
            total_words += len(text.split())
            total_chars += len(text)
            page_metrics.append(_page_metrics(page, dpi))
            d = _page_embedded_dpi(page)
            if d and 30 < d < 2400:   # sanity bounds
                dpi_samples.append(d)

        doc.close()

    if not page_metrics:
        return {}

    lap  = [m["laplacian_var"] for m in page_metrics]
    ten  = [m["tenengrad"]     for m in page_metrics]
    con  = [m["rms_contrast"]  for m in page_metrics]
    bri  = [m["brightness"]    for m in page_metrics]
    skw  = [m["skew"]          for m in page_metrics]

    if native_docs > 0 and scanned_docs > 0:
        text_method = f"mixed ({native_docs} native, {scanned_docs} scanned)"
    elif native_docs > 0:
        text_method = "native"
    else:
        text_method = "OCR/scanned"

    return {
        "laplacian_variance_blur":    round(float(np.mean(lap)), 2),
        "tenengrad_sharpness":        round(float(np.mean(ten)), 2),
        "rms_contrast":               round(float(np.mean(con)), 2),
        "mean_brightness":            round(float(np.mean(bri)), 2),
        "skew_angle_deg":             round(float(np.mean(skw)), 3),
        "resolution_dpi":             round(float(np.mean(dpi_samples))) if dpi_samples else None,
        "text_extraction_method":     text_method,
        "n_pages_extracted":          total_pages,
        "ocr_confidence_avg_pct":     None,   # requires MS Foundry OCR
        "words_per_page_avg":         round(total_words / total_pages, 1),
        "chars_per_page_avg_redacted": round(total_chars / total_pages, 1),
    }


# ── Main pipeline ─────────────────────────────────────────────────────────────

def process_sheet():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    headers    = [cell.value for cell in ws[1]]
    header_idx = {h: i for i, h in enumerate(headers) if h}
    rows_data  = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]

    folder_map = build_folder_map()

    # ── Step 1: compute derived + PDF metrics per row ─────────────────────────
    derived_list = []
    pdf_raw      = []   # raw PDF metrics for threshold computation

    for row_idx, row in enumerate(rows_data):
        surgeon   = row[header_idx["surgeon"]] or ""
        initials  = row[header_idx["patient_initials"]] or ""
        comment   = row[header_idx["comments"]] or ""

        # Comment-based quality flags
        doc_quality_flag, doc_quality_note = classify_doc_quality(comment)
        text_quality_flag = classify_text_quality(comment, doc_quality_flag)

        # Feature error counting
        n_in_source = n_not_in_source = n_ai_errors_total = 0
        n_ai_errors_minor = n_ai_errors_major = 0
        max_severity   = 1
        error_features = []
        imaging_errors = path_errors = 0

        for feat_name, src_col, _hum_col, ai_col in FEATURE_TRIPLETS:
            src_idx = col_letter_to_idx(src_col)
            ai_idx  = col_letter_to_idx(ai_col)
            src_val = row[src_idx]
            ai_val  = row[ai_idx]
            try:
                src_int = int(src_val)
            except (TypeError, ValueError):
                src_int = None

            if src_int == 1:
                n_in_source += 1
                try:
                    ai_int = int(ai_val)
                except (TypeError, ValueError):
                    ai_int = None
                if ai_int in (2, 3):
                    n_ai_errors_total += 1
                    error_features.append(feat_name)
                    max_severity = max(max_severity, ai_int)
                    if ai_int == 2:
                        n_ai_errors_minor += 1
                        (imaging_errors if feat_name in IMAGING_FEATURES else path_errors).__iadd__ if False else None
                        if feat_name in IMAGING_FEATURES:
                            imaging_errors += 1
                        else:
                            path_errors += 1
                    else:
                        n_ai_errors_major += 1
                        if feat_name in IMAGING_FEATURES:
                            imaging_errors += 1
                        else:
                            path_errors += 1
            elif src_int == 0:
                n_not_in_source += 1

        total_features = len(FEATURE_TRIPLETS)
        ai_error_rate  = (n_ai_errors_total / n_in_source) if n_in_source > 0 else 0.0
        pct_available  = (n_in_source / total_features) * 100 if total_features > 0 else 0.0

        severity_map   = {1: "none", 2: "minor", 3: "major"}
        max_sev_label  = severity_map.get(max_severity, "none")
        if n_ai_errors_total == 0:
            max_sev_label = "none"

        pdf_type     = infer_pdf_type(comment, doc_quality_flag)
        doc_type     = infer_doc_type(comment, error_features)
        quality_attr = quality_attribution(doc_quality_flag, text_quality_flag, comment)

        # ── PDF processing ─────────────────────────────────────────────────
        surg_code = surgeon_to_code(surgeon) if surgeon else ""
        folder_key = (surg_code, initials)
        folder_path = folder_map.get(folder_key)

        if folder_path:
            print(f"  Processing {os.path.basename(folder_path)} ({surg_code}_{initials}) ...")
            pdf_m = process_case_pdfs(folder_path)
        else:
            print(f"  No fab_source folder for ({surg_code}, {initials}) — skipping PDF metrics")
            pdf_m = {}

        pdf_raw.append(pdf_m)

        derived_list.append({
            # Feature availability
            "n_features_in_source":          n_in_source,
            "n_features_not_in_source":      n_not_in_source,
            "pct_features_available":        round(pct_available, 1),
            # AI errors
            "n_ai_errors_total":             n_ai_errors_total,
            "n_ai_errors_minor":             n_ai_errors_minor,
            "n_ai_errors_major":             n_ai_errors_major,
            "ai_error_rate":                 round(ai_error_rate, 3),
            "max_ai_error_severity":         max_sev_label,
            "error_features_list":           "; ".join(error_features) or "none",
            "imaging_feature_error_count":   imaging_errors,
            "pathology_feature_error_count": path_errors,
            # Comment-based doc/text quality
            "doc_type_inferred":             doc_type,
            "pdf_type_inferred":             pdf_type,
            "doc_quality_flag":              doc_quality_flag,
            "text_quality_flag":             text_quality_flag,
            "quality_error_attribution":     quality_attr,
            "doc_quality_note":              doc_quality_note,
            # PDF image quality (populated from actual PDFs below)
            "laplacian_variance_blur":       pdf_m.get("laplacian_variance_blur"),
            "tenengrad_sharpness":           pdf_m.get("tenengrad_sharpness"),
            "rms_contrast":                  pdf_m.get("rms_contrast"),
            "mean_brightness":               pdf_m.get("mean_brightness"),
            "skew_angle_deg":                pdf_m.get("skew_angle_deg"),
            "resolution_dpi":                pdf_m.get("resolution_dpi"),
            "is_blurry":                     None,   # set after threshold pass below
            "is_low_contrast":               None,
            # OCR / text extraction
            "text_extraction_method":        pdf_m.get("text_extraction_method"),
            "n_pages_extracted":             pdf_m.get("n_pages_extracted"),
            "ocr_confidence_avg_pct":        None,   # requires MS Foundry OCR
            "words_per_page_avg":            pdf_m.get("words_per_page_avg"),
            "chars_per_page_avg_redacted":   pdf_m.get("chars_per_page_avg_redacted"),
            # Performance stratum
            "performance_stratum":           (
                "poor_doc"           if doc_quality_flag == 2 else
                "possible_doc_issue" if doc_quality_flag == 1 else
                "good_doc"
            ),
        })

    # ── Step 2: compute is_blurry / is_low_contrast thresholds ───────────────
    lap_vals = [d["laplacian_variance_blur"] for d in derived_list
                if d["laplacian_variance_blur"] is not None]
    con_vals = [d["rms_contrast"] for d in derived_list
                if d["rms_contrast"] is not None]

    lap_p25 = float(np.percentile(lap_vals, 25)) if lap_vals else None
    con_p25 = float(np.percentile(con_vals, 25)) if con_vals else None

    for d in derived_list:
        lv = d["laplacian_variance_blur"]
        cv_ = d["rms_contrast"]
        d["is_blurry"]       = ("Y" if (lap_p25 is not None and lv is not None and lv <= lap_p25)
                                else ("N" if lv is not None else None))
        d["is_low_contrast"] = ("Y" if (con_p25 is not None and cv_ is not None and cv_ <= con_p25)
                                else ("N" if cv_ is not None else None))

    # ── Step 3: write new columns to sheet ───────────────────────────────────
    new_col_names = list(derived_list[0].keys())
    start_col = ws.max_column + 1

    header_fill = PatternFill("solid", fgColor="1F497D")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for i, col_name in enumerate(new_col_names):
        col_idx = start_col + i
        cell    = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    quality_fills = {
        0: PatternFill("solid", fgColor="C6EFCE"),
        1: PatternFill("solid", fgColor="FFEB9C"),
        2: PatternFill("solid", fgColor="FFC7CE"),
    }
    stratum_fills = {
        "poor_doc":           PatternFill("solid", fgColor="FFC7CE"),
        "possible_doc_issue": PatternFill("solid", fgColor="FFEB9C"),
        "good_doc":           PatternFill("solid", fgColor="C6EFCE"),
    }
    stub_fill = PatternFill("solid", fgColor="F2F2F2")
    flag_fills = {
        "Y": PatternFill("solid", fgColor="FFC7CE"),
        "N": PatternFill("solid", fgColor="C6EFCE"),
    }

    for row_offset, d in enumerate(derived_list, start=2):
        for i, col_name in enumerate(new_col_names):
            col_idx = start_col + i
            val     = d[col_name]
            cell    = ws.cell(row=row_offset, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            if val is None:
                cell.fill = stub_fill
            elif col_name == "doc_quality_flag" and isinstance(val, int):
                cell.fill = quality_fills.get(val, PatternFill())
            elif col_name == "text_quality_flag" and isinstance(val, int):
                cell.fill = quality_fills.get(val, PatternFill())
            elif col_name == "performance_stratum":
                cell.fill = stratum_fills.get(val, PatternFill())
            elif col_name == "max_ai_error_severity":
                cell.fill = quality_fills.get(
                    {"major": 2, "minor": 1}.get(val, 0), PatternFill()
                )
            elif col_name in ("is_blurry", "is_low_contrast"):
                cell.fill = flag_fills.get(val, PatternFill())

    # ── Step 4: summary block ─────────────────────────────────────────────────
    strata_counts = {"poor_doc": 0, "possible_doc_issue": 0, "good_doc": 0}
    strata_errors = {"poor_doc": [], "possible_doc_issue": [], "good_doc": []}
    strata_major  = {"poor_doc": 0, "possible_doc_issue": 0, "good_doc": 0}
    for d in derived_list:
        s = d["performance_stratum"]
        strata_counts[s] += 1
        strata_errors[s].append(d["n_ai_errors_total"])
        if d["n_ai_errors_major"] > 0:
            strata_major[s] += 1

    def mean_or_na(lst):
        return round(sum(lst) / len(lst), 2) if lst else "N/A"

    blurry_cases = [d for d in derived_list if d.get("is_blurry") == "Y"]
    lo_con_cases = [d for d in derived_list if d.get("is_low_contrast") == "Y"]

    summary_start = ws.max_row + 2
    summary_data = [
        ["DOC & TEXT EVAL PIPELINE — AI_Has_3 Fabrication Edge Cases", "", "", "", "", ""],
        ["Stratified by Document Quality (comment + PDF metrics)", "", "", "", "", ""],
        ["", "", "", "", "", ""],
        ["Performance Stratum", "N Cases", "Avg AI Errors", "Cases w/ Major Error",
         "Quality Attribution", "Blurry / Low-contrast"],
        [
            "good_doc (no quality flag)",
            strata_counts["good_doc"],
            mean_or_na(strata_errors["good_doc"]),
            strata_major["good_doc"],
            "-",
            f"{sum(1 for d in derived_list if d['performance_stratum']=='good_doc' and d.get('is_blurry')=='Y')} blurry",
        ],
        [
            "possible_doc_issue (flag=1)",
            strata_counts["possible_doc_issue"],
            mean_or_na(strata_errors["possible_doc_issue"]),
            strata_major["possible_doc_issue"],
            "Possible quality contribution",
            f"{sum(1 for d in derived_list if d['performance_stratum']=='possible_doc_issue' and d.get('is_blurry')=='Y')} blurry",
        ],
        [
            "poor_doc (flag=2, confirmed)",
            strata_counts["poor_doc"],
            mean_or_na(strata_errors["poor_doc"]),
            strata_major["poor_doc"],
            "Quality likely contributed to error",
            f"{sum(1 for d in derived_list if d['performance_stratum']=='poor_doc' and d.get('is_blurry')=='Y')} blurry",
        ],
        ["", "", "", "", "", ""],
        ["PDF IMAGE QUALITY THRESHOLDS (25th percentile within fabrication set)", "", "", "", "", ""],
        [
            f"is_blurry threshold (laplacian_var <= {lap_p25:.1f})" if lap_p25 else "is_blurry: n/a",
            f"{len(blurry_cases)} cases flagged blurry", "", "", "", "",
        ],
        [
            f"is_low_contrast threshold (rms_contrast <= {con_p25:.1f})" if con_p25 else "is_low_contrast: n/a",
            f"{len(lo_con_cases)} cases flagged low-contrast", "", "", "", "",
        ],
        ["", "", "", "", "", ""],
        ["STUB COLUMNS — Populate with MS Foundry General Documents API", "", "", "", "", ""],
        ["ocr_confidence_avg_pct",
         "Average OCR confidence from Azure Document Intelligence",
         "POST to https://...cognitiveservices.azure.com/documentintelligence/...",
         "", "", ""],
    ]

    for offset, srow in enumerate(summary_data):
        r = summary_start + offset
        for c, val in enumerate(srow, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if offset in (0, 1):
                cell.fill = PatternFill("solid", fgColor="1F497D")
                cell.font = Font(bold=True, color="FFFFFF", size=11)
            elif offset == 3:
                cell.fill = PatternFill("solid", fgColor="D9E1F2")
                cell.font = Font(bold=True, size=10)
            elif offset == 8:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
                cell.font = Font(bold=True, size=10)
            elif offset == 12:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
                cell.font = Font(bold=True, size=10)
            elif offset in (4, 5, 6):
                skey = srow[0].split("(")[0].strip().replace(" ", "_")
                for k, f in stratum_fills.items():
                    if k.startswith(skey[:7]):
                        cell.fill = f
                        break

    ws.freeze_panes = "A2"
    wb.save(EXCEL_PATH)

    # ── Console report ────────────────────────────────────────────────────────
    print()
    print("=" * 70)
    print("FAB PDF PIPELINE COMPLETE")
    print(f"  Written to: {EXCEL_PATH}")
    print(f"  New columns: {len(new_col_names)} starting at {get_column_letter(start_col)}")
    print()
    print("=== Stratified Analysis ===")
    for stratum in ["good_doc", "possible_doc_issue", "poor_doc"]:
        n   = strata_counts[stratum]
        avg = mean_or_na(strata_errors[stratum])
        maj = strata_major[stratum]
        print(f"  {stratum:<30}  N={n}  avg_ai_errors={avg}  major_error_cases={maj}")

    print()
    print(f"  is_blurry threshold:       laplacian_var <= {lap_p25:.1f}" if lap_p25 else "  is_blurry: no data")
    print(f"  is_low_contrast threshold: rms_contrast  <= {con_p25:.1f}" if con_p25 else "  is_low_contrast: no data")
    print(f"  Blurry cases: {len(blurry_cases)}  |  Low-contrast cases: {len(lo_con_cases)}")

    print()
    print("=== Per-Case Summary ===")
    for row, d in zip(rows_data, derived_list):
        surgeon_str  = str(row[header_idx["surgeon"]] or "")
        initials_str = str(row[header_idx["patient_initials"]] or "")
        scode = surgeon_to_code(surgeon_str) if surgeon_str else "?"
        folder_name = f"{scode}_{initials_str}"
        n_pages = d.get("n_pages_extracted") or "N/A"
        lap_v   = d.get("laplacian_variance_blur")
        text_m  = d.get("text_extraction_method") or "N/A"
        print(
            f"  {folder_name:<18}  errors={d['n_ai_errors_total']}"
            f" (minor={d['n_ai_errors_minor']}, major={d['n_ai_errors_major']})"
            f"  stratum={d['performance_stratum']}"
            f"  pages={n_pages}"
            f"  lap_var={lap_v}"
            f"  text={text_m}"
        )


if __name__ == "__main__":
    process_sheet()
