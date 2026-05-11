"""
Doc and Text Eval Pipeline - AI_Has_3 Sheet
Based on: LLM Prompt Engineering for performance and safety optimization
          and failure-mode analysis (Robert James, MD / Dr. Tracy Ann-Moo, MSK)

Pipeline stages (from presentation):
  1. Deidentification check and doc typing
  2. Native vs scanned PDF identification
  3. PDF image quality assessment (blur, contrast, brightness, skew, DPI)
  4. OCR vs direct text extraction
  5. Text quality metrics (words/page, chars/page, OCR confidence)
  6. Performance stratified by document and text quality

For AI_Has_3 (edge cases with >= 3 AI errors):
  - Computes all derivable metrics from existing validation data
  - Infers doc/text quality flags from comments and error patterns
  - Adds stub columns for PDF-level metrics (populated when actual PDFs are processed)
  - Adds stratified error analysis by quality tier
"""

import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH = "/Users/robertjames/Downloads/llm_validation_failure_analysis.xlsx"
SHEET_NAME = "AI_Has_3"

# Status code legend
# 0  = feature absent from source document (→ human/ai are 'NA')
# 1  = correctly extracted
# 2  = minor error
# 3  = major/clinically significant error
# NA = not applicable (feature not present in source)

FEATURE_TRIPLETS = [
    ("lesion_size",                    "G", "H", "I"),
    ("laterality",                     "J", "K", "L"),
    ("lesion_location",                "M", "N", "O"),
    ("calcifications_asymmetry",       "P", "Q", "R"),
    ("additional_enhancement_mri",     "S", "T", "U"),
    ("extent",                         "V", "W", "X"),
    ("accurate_clip_placement",        "Y", "Z", "AA"),
    ("workup_recommendation",          "AB", "AC", "AD"),
    ("lymph_node",                     "AE", "AF", "AG"),
    ("chronology_preserved",           "AH", "AI", "AJ"),
    ("biopsy_method",                  "AK", "AL", "AM"),
    ("invasive_component_size_path",   "AN", "AO", "AP"),
    ("histologic_diagnosis",           "AQ", "AR", "AS"),
    ("receptor_status",                "AT", "AU", "AV"),
]

IMAGING_FEATURES = {
    "lesion_size", "laterality", "lesion_location",
    "calcifications_asymmetry", "additional_enhancement_mri",
    "extent", "accurate_clip_placement", "workup_recommendation",
    "lymph_node", "chronology_preserved",
}
PATHOLOGY_FEATURES = {
    "biopsy_method", "invasive_component_size_path",
    "histologic_diagnosis", "receptor_status",
}

# Keywords in comments indicating document/text quality issues
QUALITY_KEYWORDS_CONFIRMED = [
    "blurry", "blur", "poor quality pdf", "poor quality", "not clear", "low quality",
    "image quality", "scanned",
]
QUALITY_KEYWORDS_POSSIBLE = [
    "uploaded text", "discrepancy", "text", "long", "truncated", "cut off",
    "missing", "difficult", "unclear", "confused", "confusion",
]
# Keywords suggesting OCR-type numeric errors
SIZE_ERROR_PATTERN = re.compile(r"\d+\.?\d*\s*(cm|mm|instead|rather|than|vs\.?)", re.IGNORECASE)


def col_letter_to_idx(letter: str) -> int:
    """Convert Excel column letter(s) to 0-based index."""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result - 1


def classify_doc_quality(comment: str) -> tuple[int, str]:
    """
    Returns (doc_quality_flag, note).
    0 = good/unknown, 1 = possible quality concern, 2 = confirmed quality issue
    """
    if not comment:
        return 0, ""
    cl = comment.lower()
    for kw in QUALITY_KEYWORDS_CONFIRMED:
        if kw in cl:
            return 2, f"Confirmed quality issue: '{kw}' mentioned in comment"
    for kw in QUALITY_KEYWORDS_POSSIBLE:
        if kw in cl:
            return 1, f"Possible quality concern: '{kw}' in comment"
    if SIZE_ERROR_PATTERN.search(comment):
        return 1, "Possible OCR/text extraction error: numeric size discrepancy"
    return 0, ""


def classify_text_quality(comment: str, doc_quality_flag: int) -> int:
    """
    0 = good, 1 = possible text issue, 2 = confirmed text/OCR issue.
    Numeric transpositions (e.g. 1.4 vs 0.4) suggest OCR errors.
    """
    if not comment:
        return doc_quality_flag  # inherit from doc quality
    cl = comment.lower()
    numeric_swap = re.search(
        r"(\d+\.?\d*)\s*(instead of|rather than|was|vs\.?)\s*(\d+\.?\d*)", cl
    )
    if numeric_swap:
        return max(doc_quality_flag, 1)
    return doc_quality_flag


def infer_pdf_type(comment: str, doc_quality_flag: int) -> str:
    """Infer native vs scanned from context."""
    if not comment:
        return "unknown"
    cl = comment.lower()
    if any(k in cl for k in ["blurry", "poor quality pdf", "scanned", "image quality"]):
        return "scanned"
    if any(k in cl for k in ["selectable", "native", "word", "export"]):
        return "native"
    if doc_quality_flag == 2:
        return "scanned"
    return "unknown"


def infer_doc_type(comment: str, feature_errors: list[str]) -> str:
    """
    Infer document type from error context.
    Both imaging and pathology features present → mixed.
    """
    has_imaging = any(f in IMAGING_FEATURES for f in feature_errors)
    has_path = any(f in PATHOLOGY_FEATURES for f in feature_errors)
    if not feature_errors:
        return "mixed"  # breast cancer workup always has both
    if has_imaging and has_path:
        return "mixed"
    if has_imaging:
        return "radiology"
    if has_path:
        return "pathology"
    return "mixed"


def quality_attribution(doc_quality_flag: int, text_quality_flag: int,
                         comment: str) -> str:
    """Assess whether quality likely contributed to error."""
    if doc_quality_flag == 2 or text_quality_flag == 2:
        return "Confirmed"
    if doc_quality_flag == 1 or text_quality_flag == 1:
        return "Possible"
    if comment and SIZE_ERROR_PATTERN.search(comment):
        return "Possible"
    return "No"


def process_sheet():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    # ------------------------------------------------------------------ #
    # 1. Read all data rows into dicts
    # ------------------------------------------------------------------ #
    headers = [cell.value for cell in ws[1]]
    header_idx = {h: i for i, h in enumerate(headers) if h}

    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows_data.append(list(row))

    # ------------------------------------------------------------------ #
    # 2. Compute derived metrics per row
    # ------------------------------------------------------------------ #
    derived = []
    for row in rows_data:
        comment = row[header_idx["comments"]] or ""

        n_in_source = 0
        n_not_in_source = 0
        n_ai_errors_total = 0
        n_ai_errors_minor = 0
        n_ai_errors_major = 0
        max_severity = 1
        error_features = []
        imaging_errors = 0
        path_errors = 0

        for feat_name, src_col, hum_col, ai_col in FEATURE_TRIPLETS:
            src_idx = col_letter_to_idx(src_col)
            ai_idx  = col_letter_to_idx(ai_col)

            src_val = row[src_idx]
            ai_val  = row[ai_idx]

            try:
                src_int = int(src_val)
            except (TypeError, ValueError):
                src_int = None

            if src_int == 1:
                n_in_source += 1
                try:
                    ai_int = int(ai_val)
                except (TypeError, ValueError):
                    ai_int = None
                if ai_int in (2, 3):
                    n_ai_errors_total += 1
                    error_features.append(feat_name)
                    max_severity = max(max_severity, ai_int)
                    if ai_int == 2:
                        n_ai_errors_minor += 1
                        if feat_name in IMAGING_FEATURES:
                            imaging_errors += 1
                        else:
                            path_errors += 1
                    elif ai_int == 3:
                        n_ai_errors_major += 1
                        if feat_name in IMAGING_FEATURES:
                            imaging_errors += 1
                        else:
                            path_errors += 1
            elif src_int == 0:
                n_not_in_source += 1

        total_features = len(FEATURE_TRIPLETS)
        ai_error_rate = (n_ai_errors_total / n_in_source) if n_in_source > 0 else 0.0
        pct_available = (n_in_source / total_features) * 100 if total_features > 0 else 0.0

        error_features_str = "; ".join(error_features) if error_features else "none"

        doc_quality_flag, doc_quality_note = classify_doc_quality(comment)
        text_quality_flag = classify_text_quality(comment, doc_quality_flag)
        pdf_type = infer_pdf_type(comment, doc_quality_flag)
        doc_type = infer_doc_type(comment, error_features)
        quality_attr = quality_attribution(doc_quality_flag, text_quality_flag, comment)

        # Severity label
        severity_map = {1: "none", 2: "minor", 3: "major"}
        max_sev_label = severity_map.get(max_severity, "none")
        if n_ai_errors_total == 0:
            max_sev_label = "none"

        derived.append({
            # ---- Feature availability metrics ----
            "n_features_in_source":         n_in_source,
            "n_features_not_in_source":     n_not_in_source,
            "pct_features_available":       round(pct_available, 1),
            # ---- AI error metrics ----
            "n_ai_errors_total":            n_ai_errors_total,
            "n_ai_errors_minor":            n_ai_errors_minor,
            "n_ai_errors_major":            n_ai_errors_major,
            "ai_error_rate":                round(ai_error_rate, 3),
            "max_ai_error_severity":        max_sev_label,
            "error_features_list":          error_features_str,
            "imaging_feature_error_count":  imaging_errors,
            "pathology_feature_error_count": path_errors,
            # ---- Doc/text quality assessment ----
            "doc_type_inferred":            doc_type,
            "pdf_type_inferred":            pdf_type,
            "doc_quality_flag":             doc_quality_flag,
            "text_quality_flag":            text_quality_flag,
            "quality_error_attribution":    quality_attr,
            "doc_quality_note":             doc_quality_note,
            # ---- PDF image quality stubs (populate via PyMuPDF + OpenCV) ----
            "laplacian_variance_blur":      None,  # Variance of Laplacian; low = blurry
            "tenengrad_sharpness":          None,  # Tenengrad gradient energy
            "rms_contrast":                 None,  # Intensity spread p95-p5
            "mean_brightness":              None,  # Mean pixel brightness
            "skew_angle_deg":               None,  # Detected skew angle
            "resolution_dpi":               None,  # Page DPI
            "is_blurry":                    None,  # Y/N (bottom 10th %ile of laplacian_variance)
            "is_low_contrast":              None,  # Y/N (bottom 10th %ile of rms_contrast)
            # ---- OCR / text extraction stubs (populate via MS Foundry OCR or PyMuPDF) ----
            "text_extraction_method":       None,  # direct_transcription / OCR
            "n_pages_extracted":            None,
            "ocr_confidence_avg_pct":       None,
            "words_per_page_avg":           None,
            "chars_per_page_avg_redacted":  None,
            # ---- Stratified performance tier ----
            "performance_stratum":          (
                "poor_doc"    if doc_quality_flag == 2 else
                "possible_doc_issue" if doc_quality_flag == 1 else
                "good_doc"
            ),
        })

    # ------------------------------------------------------------------ #
    # 3. Append new header columns starting at AX (col 50)
    # ------------------------------------------------------------------ #
    new_col_names = list(derived[0].keys())
    start_col = ws.max_column + 1  # = 50 (AX)

    # Header row styling
    header_fill   = PatternFill("solid", fgColor="1F497D")
    header_font   = Font(bold=True, color="FFFFFF", size=10)
    section_fills = {
        "feature_avail":  PatternFill("solid", fgColor="D9E1F2"),
        "ai_error":       PatternFill("solid", fgColor="FCE4D6"),
        "doc_quality":    PatternFill("solid", fgColor="E2EFDA"),
        "pdf_img_stub":   PatternFill("solid", fgColor="FFF2CC"),
        "ocr_stub":       PatternFill("solid", fgColor="EDEDED"),
        "stratum":        PatternFill("solid", fgColor="D6DCE4"),
    }

    def col_section(col_name: str) -> str:
        if col_name.startswith("n_features") or col_name.startswith("pct_"):
            return "feature_avail"
        if col_name.startswith("n_ai") or col_name in (
            "ai_error_rate", "max_ai_error_severity", "error_features_list",
            "imaging_feature_error_count", "pathology_feature_error_count"
        ):
            return "ai_error"
        if col_name in ("doc_type_inferred", "pdf_type_inferred",
                         "doc_quality_flag", "text_quality_flag",
                         "quality_error_attribution", "doc_quality_note"):
            return "doc_quality"
        if col_name in ("laplacian_variance_blur", "tenengrad_sharpness",
                         "rms_contrast", "mean_brightness", "skew_angle_deg",
                         "resolution_dpi", "is_blurry", "is_low_contrast"):
            return "pdf_img_stub"
        if col_name in ("text_extraction_method", "n_pages_extracted",
                         "ocr_confidence_avg_pct", "words_per_page_avg",
                         "chars_per_page_avg_redacted"):
            return "ocr_stub"
        return "stratum"

    for i, col_name in enumerate(new_col_names):
        col_idx = start_col + i
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # ------------------------------------------------------------------ #
    # 4. Write derived values into data rows
    # ------------------------------------------------------------------ #
    quality_fills = {
        0: PatternFill("solid", fgColor="C6EFCE"),   # green = good/unknown
        1: PatternFill("solid", fgColor="FFEB9C"),   # yellow = possible
        2: PatternFill("solid", fgColor="FFC7CE"),   # red = confirmed issue
    }
    stratum_fills = {
        "poor_doc":           PatternFill("solid", fgColor="FFC7CE"),
        "possible_doc_issue": PatternFill("solid", fgColor="FFEB9C"),
        "good_doc":           PatternFill("solid", fgColor="C6EFCE"),
    }

    for row_idx, d in enumerate(derived, start=2):
        for i, col_name in enumerate(new_col_names):
            col_idx = start_col + i
            val = d[col_name]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            if col_name == "doc_quality_flag" and isinstance(val, int):
                cell.fill = quality_fills.get(val, PatternFill())
            elif col_name == "text_quality_flag" and isinstance(val, int):
                cell.fill = quality_fills.get(val, PatternFill())
            elif col_name == "performance_stratum":
                cell.fill = stratum_fills.get(val, PatternFill())
            elif col_name == "max_ai_error_severity":
                if val == "major":
                    cell.fill = quality_fills[2]
                elif val == "minor":
                    cell.fill = quality_fills[1]
                else:
                    cell.fill = quality_fills[0]
            # Shade stub columns
            elif d[col_name] is None:
                cell.fill = PatternFill("solid", fgColor="F2F2F2")

    # ------------------------------------------------------------------ #
    # 5. Add a summary block below data rows (Part 3 stratified analysis)
    # ------------------------------------------------------------------ #
    summary_start_row = ws.max_row + 2
    section_header_font = Font(bold=True, size=10)
    section_header_fill = PatternFill("solid", fgColor="1F497D")

    # Count rows per stratum
    strata_counts = {"poor_doc": 0, "possible_doc_issue": 0, "good_doc": 0}
    strata_errors = {"poor_doc": [], "possible_doc_issue": [], "good_doc": []}
    strata_major  = {"poor_doc": 0, "possible_doc_issue": 0, "good_doc": 0}
    for d in derived:
        s = d["performance_stratum"]
        strata_counts[s] += 1
        strata_errors[s].append(d["n_ai_errors_total"])
        if d["n_ai_errors_major"] > 0:
            strata_major[s] += 1

    def mean_or_na(lst):
        return round(sum(lst) / len(lst), 2) if lst else "N/A"

    summary_data = [
        ["DOC & TEXT EVAL PIPELINE SUMMARY — AI_Has_3 (Edge Cases)", "", "", "", ""],
        ["Stratified by Document Quality (Part 3 Analysis)", "", "", "", ""],
        ["", "", "", "", ""],
        ["Performance Stratum", "N Cases", "Avg AI Errors", "Cases w/ Major Error", "Quality Attribution"],
        [
            "good_doc (no quality flag)",
            strata_counts["good_doc"],
            mean_or_na(strata_errors["good_doc"]),
            strata_major["good_doc"],
            "—",
        ],
        [
            "possible_doc_issue (flag=1)",
            strata_counts["possible_doc_issue"],
            mean_or_na(strata_errors["possible_doc_issue"]),
            strata_major["possible_doc_issue"],
            "Possible quality contribution",
        ],
        [
            "poor_doc (flag=2, confirmed)",
            strata_counts["poor_doc"],
            mean_or_na(strata_errors["poor_doc"]),
            strata_major["poor_doc"],
            "Quality likely contributed to error",
        ],
        ["", "", "", "", ""],
        ["STUB COLUMNS — Populate with PyMuPDF + OpenCV + MS Foundry OCR", "", "", "", ""],
        [
            "laplacian_variance_blur",
            "Variance of Laplacian (low = blurry)",
            "PyMuPDF → page.get_pixmap() → cv2.Laplacian(img, cv2.CV_64F).var()",
            "", "",
        ],
        [
            "tenengrad_sharpness",
            "Gradient energy (Sobel)",
            "cv2.Sobel on page image; sum of squared gradients",
            "", "",
        ],
        [
            "rms_contrast",
            "Intensity spread p95-p5",
            "np.percentile(img, 95) - np.percentile(img, 5)",
            "", "",
        ],
        [
            "mean_brightness",
            "Mean pixel intensity (0-255)",
            "np.mean(img)",
            "", "",
        ],
        [
            "skew_angle_deg",
            "Page skew (degrees)",
            "cv2.minAreaRect on thresholded text contours",
            "", "",
        ],
        [
            "resolution_dpi",
            "Page DPI",
            "Derived from PDF metadata or pixmap matrix",
            "", "",
        ],
        [
            "is_blurry / is_low_contrast",
            "Y/N threshold flags",
            "Bottom 10th percentile within dataset",
            "", "",
        ],
        [
            "ocr_confidence_avg_pct",
            "Average OCR confidence",
            "MS Foundry General Documents API confidence scores",
            "", "",
        ],
        [
            "words_per_page_avg / chars_per_page_avg_redacted",
            "Text density after redaction",
            "Token count from extracted text after PHI redaction",
            "", "",
        ],
    ]

    for offset, srow in enumerate(summary_data):
        r = summary_start_row + offset
        for c, val in enumerate(srow, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if offset in (0, 1):
                cell.fill = PatternFill("solid", fgColor="1F497D")
                cell.font = Font(bold=True, color="FFFFFF", size=11)
            elif offset == 3:
                cell.fill = PatternFill("solid", fgColor="D9E1F2")
                cell.font = Font(bold=True, size=10)
            elif offset == 8:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
                cell.font = Font(bold=True, size=10)
            elif offset in range(4, 8):
                stratum_key = srow[0].split(" ")[0] if srow[0] else ""
                if stratum_key == "good_doc":
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                elif stratum_key == "possible_doc_issue":
                    cell.fill = PatternFill("solid", fgColor="FFEB9C")
                elif stratum_key == "poor_doc":
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")

    # Freeze panes at row 2
    ws.freeze_panes = "A2"

    wb.save(EXCEL_PATH)
    print(f"Pipeline columns written to '{SHEET_NAME}' in {EXCEL_PATH}")
    print(f"Added {len(new_col_names)} new columns starting at column {get_column_letter(start_col)}")
    print()
    print("=== Stratified Analysis (Part 3) ===")
    for stratum in ["good_doc", "possible_doc_issue", "poor_doc"]:
        n = strata_counts[stratum]
        avg_err = mean_or_na(strata_errors[stratum])
        maj = strata_major[stratum]
        print(f"  {stratum:30s}  N={n}  avg_ai_errors={avg_err}  cases_w_major_error={maj}")
    print()
    print("=== Per-Case Summary ===")
    for i, (row, d) in enumerate(zip(rows_data, derived), start=1):
        initials = row[3]
        print(
            f"  Case {i:2d} ({initials}): "
            f"errors={d['n_ai_errors_total']} "
            f"(minor={d['n_ai_errors_minor']}, major={d['n_ai_errors_major']}) | "
            f"stratum={d['performance_stratum']} | "
            f"quality_attr={d['quality_error_attribution']}"
        )


if __name__ == "__main__":
    process_sheet()
